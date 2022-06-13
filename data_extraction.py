import openpyxl as op
import re
import pandas as pd
import datetime
import logging
import pathlib as pl
import json
import dateutil.parser

# region Logging
LOGFOLDER = pl.Path('Logs')
FILENAME = f'{pl.Path(__file__).stem}.log'
FORMAT = "%(asctime)s - %(message)s"
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(logging.StreamHandler())
if not LOGFOLDER.exists():
    LOGFOLDER.mkdir()
logfile = LOGFOLDER / FILENAME
handler = logging.FileHandler(logfile, mode='w')
formatter = logging.Formatter(FORMAT)
handler.setFormatter(formatter)
logger.addHandler(handler)


# endregion


# region Parameters
with open('programs.json', 'r') as f:
    programs = json.load(f)
    programs = [key for key, val in programs.items() if val]
# endregion


# region Methods


def detect_line_item(cell):
    line_item_regex = "[1-9][0-9]*-[0-9]{2}.?"
    if not cell.value:
        return False
    elif cell.value in (999, '00999'):
        logger.debug(f'{cell.value=}')
        logger.debug('00999-01')
        return ['00999-01']
    else:
        logger.debug(f'{cell.value=}')
        logger.debug(re.match(line_item_regex, str(cell.value)))
        return re.match(line_item_regex, str(cell.value))


def extract_quarter(value: str, try_on_regex_fail: bool = True):
    if type(value) is not str:
        return value
    logger.debug(value)
    date_string = re.search('[0-9/]{5,}', value)
    if date_string:
        date = dateutil.parser.parse(date_string[0], dayfirst=False, yearfirst=False)
        logger.debug(date)
        return date
    if not try_on_regex_fail:
        return None
    date = dateutil.parser.parse(value, dayfirst=False, yearfirst=False, fuzzy=True)
    logger.debug(date)
    return date


def find_info_row(sheet, start_row, info_col):
    logger.debug(f'{start_row=}')
    col = sheet[info_col]
    rows = [cell.row for cell in col if cell.row >= start_row and cell.value]
    logger.debug(f'{rows=}')
    return min(rows)


def rbha_sheet_reformat(sheet):
    items = sheet['A']
    item_rows = [item.row for item in items if detect_line_item(item)]
    sub_items = sheet['B']
    for sub_item in sub_items:
        if sub_item.value:
            if len(sub_item.value) == 1:
                item_row = max(row for row in item_rows if row < sub_item.row)
                logger.debug(f'Getting sub-item from {sub_item.row}. Item from {item_row}')
                line_item = sheet[f'A{item_row}'].value
                logger.debug(f'{line_item=}\t{sub_item.value=}')
                result = line_item + sub_item.value
                sheet[f'A{sub_item.row}'] = result
    return sheet


def extract_revenues_and_expenses(filename, info_row, info_column, name_cell, quarter_cell, sheet_names,
                                  first_column, program,
                                  line_items, **kwargs):
    logger.info(pl.Path(filename).stem)
    wb = op.load_workbook(filename, data_only=True)

    data = []
    name = None
    for sheet_name in sheet_names:
        logger.info(f'\t{sheet_name}')
        try:
            sheet = wb[sheet_name]
        except KeyError as ke:
            logger.warning(f"{filename}\t{sheet_name} didn't exist")
            continue
        if program == 'RBHA':
            sheet = rbha_sheet_reformat(sheet)
        if program == 'EPD':
            info_row = find_info_row(sheet, info_row, info_column[0])
        name = str(sheet[name_cell].value).strip()
        logger.info(f'{name=}')
        quarter = extract_quarter(sheet[quarter_cell].value)
        ffy = (quarter + datetime.timedelta(days=31)).year
        data_rows = [(cell.row, li[0]) for cell in sheet[info_column[0]]
                     if (li := detect_line_item(cell))
                     ]
        data_columns = [cell.column for cell in sheet[info_row]
                        if cell.value
                        and cell.column >= first_column
                        and 'total' not in str(cell.value).lower()
                        and 'ytd' not in str(cell.value).lower()  # to handle CHP - doesn't work
                        ]
        if program == 'CHP':
            data_columns = data_columns[:kwargs['num_columns']]
        for j in data_columns:
            for i, line_item in data_rows:
                if (cell := sheet.cell(i, j)).value or cell.value == 0:
                    value = cell.value
                    if type(value) is str:
                        value = value.strip()
                        if value == '-':
                            value = 0
                    data.append(
                        {
                            'Name': name,
                            'Value': value,
                            'Quarter': quarter,
                            'File Name': pl.Path(filename).stem,
                            'FFY': ffy,
                            'Sheet': sheet_name,
                            'Column': sheet.cell(info_row, j).value,
                            'Line Item': line_item,
                        })
    df = pd.DataFrame.from_records(data)
    df = df.join(line_items, on="Line Item")
    df['Line Lookup'] = df['Line Item'].astype(str) + ' - ' + df['Line Name']
    is_not_total = df['Line Name'].apply(lambda x: 'total' not in str(x).lower())
    df = df[is_not_total]
    return df, name, filename, quarter


def run_program(program, filenames):
    output_directory = pl.Path(f'Output/{program}')
    output_directory.mkdir(parents=True, exist_ok=True)
    logger.info(f'{output_directory=}')

    with pl.Path(f'formats/{program}.json').open('r') as f:
        params = json.load(f)
    line_items = pd.read_csv(pl.Path("Line_Items") / f'{program}.csv', index_col=0)

    dfs = []
    for filename in filenames:
        file = pl.Path(filename)
        try:
            new_file = pl.Path(f'Input/{program}/{file.name}')
            file = file.rename(new_file)
            dfs.append(extract_revenues_and_expenses(file, line_items=line_items, **params))
        except FileExistsError:
            logger.warning(f'File Exists: {file}')
        except Exception as exception:
            logger.exception(file)
            logger.exception(exception)
    logger.info("Done processing data")
    destinations = []
    for matching_df, df_name, filename, quarter in dfs:
        if program == "CHP":
            matching_df['Source'] = matching_df['Quarter']
            matching_df['Quarter'] = matching_df['Column'].apply(extract_quarter, args=(False,))
        quarter_str = quarter.strftime("%Y%m")
        destination = output_directory / f'{df_name} - {quarter_str}.csv'
        destinations.append(destination)
        matching_df.to_csv(destination,
                           index=False)
        logger.info(f'Processed {quarter_str} - {df_name:<75}{pl.Path(filename).stem}')
    for d in set(destinations):
        destinations.remove(d)
    for d in destinations:
        logger.warning(f"Duplicate found: {d}")


# endregion


def main():
    for program in programs:
        input_folder = pl.Path('Input') / program
        run_program(program, input_folder.iterdir())


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logger.exception(e)
        input()
        raise e
