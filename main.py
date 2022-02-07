import openpyxl as op
import re
import sys
import pandas as pd
import datetime
import logging
import pathlib as pl
import json
import dateutil.parser

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(logging.StreamHandler())


def detect_line_item(cell):
    line_item_regex = "[1-9][0-9]*-[0-9]{2}.?"
    if not cell.value:
        return False
    else:
        logger.debug(cell.value)
        logger.debug(re.match(line_item_regex, str(cell.value)))
        return re.match(line_item_regex, str(cell.value))


def extract_quarter(value: str):
    logger.debug(value)
    date_string = re.search('[0-9/]{5,}', value)
    if date_string:
        date = dateutil.parser.parse(date_string[0], dayfirst=False, yearfirst=False)
        logger.debug(date)
        return date
    date = dateutil.parser.parse(value, dayfirst=False, yearfirst=False, fuzzy=True)
    logger.debug(date)
    return date


def rbha_sheet_reformat(sheet):
    items = sheet['A']
    item_rows = [item.row for item in items if detect_line_item(item)]
    sub_items = sheet['B']
    for sub_item in sub_items:
        if sub_item.value:
            if len(sub_item.value) == 1:
                item_row = max(row for row in item_rows if row < sub_item.row)
                logger.debug(f'Getting sub-item from {sub_item.row}. Item from {item_row}')
                line_item = sheet[f'A{item_row}'].value
                logger.debug(f'{line_item=}\t{sub_item.value=}')
                result = line_item + sub_item.value
                sheet[f'A{sub_item.row}'] = result
    return sheet


def extract_revenues_and_expenses(filename, columns, info_row, info_column, name_cell, quarter_cell, sheet_names,
                                  first_column, program,
                                  line_items):
    logger.info(pl.Path(filename).stem)
    wb = op.load_workbook(filename, data_only=True)

    data = []
    name = None
    for sheet_name in sheet_names:
        logger.info(f'\t{sheet_name}')
        try:
            sheet = wb[sheet_name]
        except KeyError as ke:
            logger.warning(f"{filename}\t{sheet_name} didn't exist")
            continue
        if program == 'RBHA':
            sheet = rbha_sheet_reformat(sheet)
        name = sheet[name_cell].value
        logger.info(f'{name=}')
        quarter = extract_quarter(sheet[quarter_cell].value)
        data_rows = [cell.row for cell in sheet[info_column[0]]
                     if detect_line_item(cell)
                     ]
        data_columns = [cell.column for cell in sheet[info_row]
                        if cell.value
                        and cell.column >= first_column
                        and 'total' not in str(cell.value).lower()
                        ]
        for j in data_columns:
            for i in data_rows:
                if (cell := sheet.cell(i, j)).value or cell.value == 0:
                    data.append(
                        {
                            'Name': name,
                            'Value': cell.value,
                            'Quarter': quarter,
                            'Sheet': sheet_name,
                            'Column': sheet.cell(info_row, j).value,
                            'Line Item': sheet.cell(i, info_column[1]).value,
                        })
    df = pd.DataFrame.from_records(data)
    df = df.join(line_items, on="Line Item")
    df['Line Lookup'] = df['Line Item'] + ' - ' + df['Line Name']
    return df, name


def main():
    options = ['ACC', 'RBHA', 'ALTCS']
    program = None
    while program not in options:
        program = input(f"Please choose a program. Your options are: {', '.join(options)}\n\t")
    assert program in options
    now = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
    output_directory = pl.Path(f'Output/{program}/{now}')
    output_directory.mkdir(parents=True)
    handler = logging.FileHandler(output_directory / 'app.log', mode='w')
    logger.addHandler(handler)
    logger.info(f'{output_directory=}')

    with pl.Path(f'formats/{program}.json').open('r') as f:
        params = json.load(f)
    line_items = pd.read_csv(pl.Path("Line_Items") / f'{program}.csv', index_col=0)
    filenames = sys.argv[1:]
    dfs = []
    for filename in filenames:
        try:
            dfs.append(extract_revenues_and_expenses(filename, line_items=line_items, **params))
        except Exception as exception:
            logger.exception(filename)
            logger.exception(exception)
    logging.info("Done processing data")
    total_df = pd.DataFrame(columns=params['columns'])
    names = set([x[1] for x in dfs])
    for name in names:
        matching_dfs = [x[0] for x in dfs if x[1] == name]
        result = pd.DataFrame(columns=params['columns'])
        for matching_df in matching_dfs:
            result = result.append(matching_df)
        result = result.drop_duplicates()
        result.to_excel(output_directory / f'{name}.xlsx', index=False)
        total_df = total_df.append(result)
    total_df.to_excel(output_directory / 'Total.xlsx',
                      index=False)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logger.exception(e)
        input()
        raise e
